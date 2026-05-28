"""Workbook and financial statement validation."""

from __future__ import annotations

from math import isclose
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .calculations import CALCULATION_FUNCTIONS, check_balance_sheet_balances, check_cash_flow_reconciliation
from .field_codes import (
    INPUT_HEADERS,
    LINE_INPUT,
    OPTIONAL,
    PROFILE_FIELDS,
    REQUIRED,
    REQUIRED_SHEETS,
    SHEET_COMPREHENSIVE_INPUT,
    SHEET_ESSENTIAL_INPUT,
    STATEMENT_BALANCE_SHEET,
    STATEMENT_CASH_FLOW,
    STATEMENT_PROFIT_LOSS,
    TEMPLATE_COMPREHENSIVE,
    TEMPLATE_ESSENTIAL,
    get_allowed_field_codes,
    get_fields,
)
from .ingestion import dataframe_to_period_maps, detect_template_level, get_period_columns, read_company_profile, read_financial_data
from .scale import is_valid_scale
from .schema import ValidationIssue


def _issue(severity: str, message: str, **kwargs: Any) -> ValidationIssue:
    return ValidationIssue(severity=severity, message=message, **kwargs)


def _raw_numeric_issues(path: str | Path, sheet_name: str, template_level: str) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    periods = [header for header in headers if header not in INPUT_HEADERS and header is not None]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        for period in periods:
            value = record.get(period)
            if value in (None, ""):
                continue
            try:
                float(value)
            except (TypeError, ValueError):
                issues.append(
                    _issue(
                        "Error",
                        "Financial value must be numeric.",
                        template_level=template_level,
                        statement=record.get("Statement"),
                        section=record.get("Section"),
                        field_code=record.get("Field Code"),
                        period=str(period),
                        actual_value=str(value),
                    )
                )
    return issues


def _validate_required_sheets(path: str | Path) -> list[ValidationIssue]:
    wb = load_workbook(path, read_only=True)
    existing = set(wb.sheetnames)
    return [
        _issue("Error", f"Required workbook sheet is missing: {sheet}", expected_value=sheet)
        for sheet in REQUIRED_SHEETS
        if sheet not in existing
    ]


def _validate_profile(profile: dict[str, Any], template_level: str | None) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for field in PROFILE_FIELDS:
        if profile.get(field) in (None, ""):
            issues.append(_issue("Error", f"Company profile field is required: {field}", template_level=template_level, field_code=field))
    scale = profile.get("Reporting Scale")
    if scale and not is_valid_scale(str(scale)):
        issues.append(_issue("Error", "Reporting Scale is invalid.", template_level=template_level, field_code="Reporting Scale", expected_value="Units, Thousands, Millions, or Billions", actual_value=str(scale)))
    return issues


def _validate_periods(path: str | Path) -> list[ValidationIssue]:
    essential = get_period_columns(path, SHEET_ESSENTIAL_INPUT)
    comprehensive = get_period_columns(path, SHEET_COMPREHENSIVE_INPUT)
    if essential != comprehensive:
        return [_issue("Error", "Period columns must be consistent between Essential_Input and Comprehensive_Input.", expected_value=", ".join(essential), actual_value=", ".join(comprehensive))]
    if not essential:
        return [_issue("Error", "At least one period column is required.")]
    return []


def _validate_field_codes(path: str | Path, sheet_name: str, template_level: str) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    allowed = get_allowed_field_codes(template_level)
    present = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        code = record.get("Field Code")
        if not code:
            continue
        present.add(code)
        if code not in allowed:
            issues.append(_issue("Error", "Field code is not approved for this template.", template_level=template_level, statement=record.get("Statement"), section=record.get("Section"), field_code=str(code), actual_value=str(code)))
    for field in get_fields(template_level):
        if field.field_code not in present:
            severity = "Error" if field.required_level == REQUIRED else "Warning"
            issues.append(_issue(severity, "Expected field code row is missing.", template_level=template_level, statement=field.statement, section=field.section, field_code=field.field_code, expected_value=field.field_code))
    return issues


def _validate_required_values(df: pd.DataFrame, template_level: str) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    if df.empty:
        issues.append(_issue("Error", "Neither input sheet has sufficient required data.", template_level=template_level))
        return issues
    for field in get_fields(template_level):
        if field.line_type != LINE_INPUT:
            continue
        field_df = df[df["field_code"] == field.field_code]
        for _, row in field_df.iterrows():
            missing = pd.isna(row["normalized_value"])
            if missing and field.required_level == REQUIRED:
                issues.append(_issue("Error", "Required input value is missing.", template_level=template_level, statement=field.statement, section=field.section, field_code=field.field_code, period=str(row["period"])))
            elif missing and field.required_level == OPTIONAL:
                issues.append(_issue("Warning", "Optional Comprehensive field is missing; formulas may treat it as zero where applicable.", template_level=template_level, statement=field.statement, section=field.section, field_code=field.field_code, period=str(row["period"])))
    return issues


def build_calculated_period_maps(df: pd.DataFrame, template_level: str) -> dict[str, dict[str, float]]:
    maps = dataframe_to_period_maps(df)
    for period, data in maps.items():
        for field in get_fields(template_level):
            func = CALCULATION_FUNCTIONS.get(field.field_code)
            if not func:
                continue
            result = func(data, template_level)
            if result.status in ("Calculated", "Warning") and result.calculated_value is not None:
                data.setdefault(field.field_code, result.calculated_value)
    return maps


def _formula_validation(df: pd.DataFrame, template_level: str, tolerance: float = 1.0) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    maps = build_calculated_period_maps(df, template_level)
    metadata = {field.field_code: field for field in get_fields(template_level)}
    actual_lookup = {
        (str(row["period"]), row["field_code"]): row["normalized_value"]
        for _, row in df.iterrows()
        if pd.notna(row["normalized_value"])
    }
    for period, data in maps.items():
        for field in get_fields(template_level):
            func = CALCULATION_FUNCTIONS.get(field.field_code)
            if not func:
                continue
            result = func(data, template_level)
            actual = actual_lookup.get((period, field.field_code))
            if actual is None or result.calculated_value is None:
                continue
            diff = float(actual) - float(result.calculated_value)
            if not isclose(diff, 0.0, abs_tol=tolerance):
                statement = field.statement
                issues.append(
                    _issue(
                        "Error",
                        f"{statement} formula does not reconcile for {field.field_code}.",
                        template_level=template_level,
                        statement=statement,
                        section=field.section,
                        field_code=field.field_code,
                        period=period,
                        expected_value=result.calculated_value,
                        actual_value=float(actual),
                        difference=diff,
                    )
                )
        bs = check_balance_sheet_balances(data, template_level, tolerance)
        if bs.status == "Warning":
            field = metadata.get("total_assets")
            issues.append(_issue("Error", "Balance Sheet does not balance.", template_level=template_level, statement=STATEMENT_BALANCE_SHEET, section=field.section if field else None, field_code="total_assets", period=period, expected_value=data.get("total_liabilities_and_equity"), actual_value=data.get("total_assets"), difference=bs.calculated_value))
        cf = check_cash_flow_reconciliation(data, template_level, tolerance)
        if cf.status == "Warning":
            issues.append(_issue("Error", "Cash Flow reconciliation mismatch.", template_level=template_level, statement=STATEMENT_CASH_FLOW, field_code="closing_cash_balance", period=period, expected_value=(data.get("opening_cash_balance") or 0) + (data.get("net_change_in_cash") or 0), actual_value=data.get("closing_cash_balance"), difference=cf.calculated_value))
    return issues


def validate_workbook(path: str | Path) -> tuple[list[ValidationIssue], str | None]:
    issues = _validate_required_sheets(path)
    if issues:
        return issues, None
    profile = read_company_profile(path)
    template_level = detect_template_level(path, profile)
    issues.extend(_validate_profile(profile, template_level))
    issues.extend(_validate_periods(path))
    issues.extend(_validate_field_codes(path, SHEET_ESSENTIAL_INPUT, TEMPLATE_ESSENTIAL))
    issues.extend(_validate_field_codes(path, SHEET_COMPREHENSIVE_INPUT, TEMPLATE_COMPREHENSIVE))
    issues.extend(_raw_numeric_issues(path, SHEET_ESSENTIAL_INPUT, TEMPLATE_ESSENTIAL))
    issues.extend(_raw_numeric_issues(path, SHEET_COMPREHENSIVE_INPUT, TEMPLATE_COMPREHENSIVE))
    if template_level is None:
        issues.append(_issue("Error", "Neither input sheet has sufficient required data. Complete Essential_Input or Comprehensive_Input."))
        return issues, None
    _, df, _ = read_financial_data(path, template_level)
    issues.extend(_validate_required_values(df, template_level))
    issues.extend(_formula_validation(df, template_level))
    if not any(issue.severity == "Error" for issue in issues):
        issues.append(_issue("Info", "Workbook validation completed without errors.", template_level=template_level))
    return issues, template_level
