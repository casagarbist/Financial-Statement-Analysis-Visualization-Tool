"""Read standardized Excel workbooks into tidy normalized data."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .field_codes import (
    INPUT_HEADERS,
    PROFILE_FIELDS,
    REQUIRED,
    SHEET_COMPANY_PROFILE,
    SHEET_COMPREHENSIVE_INPUT,
    SHEET_ESSENTIAL_INPUT,
    TEMPLATE_COMPREHENSIVE,
    TEMPLATE_ESSENTIAL,
    get_fields,
)
from .scale import normalize_value


def read_company_profile(path: str | Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET_COMPANY_PROFILE]
    profile: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in PROFILE_FIELDS:
            profile[str(row[0])] = row[1]
    return profile


def _read_input_sheet(path: str | Path, sheet_name: str, template_level: str, profile: dict[str, Any]) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=True, read_only=True)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    period_columns = [header for header in headers if header not in INPUT_HEADERS and header is not None]
    rows: list[dict[str, Any]] = []
    scale = str(profile.get("Reporting Scale") or "Units")
    for values in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        if not record.get("Field Code"):
            continue
        for period in period_columns:
            value = record.get(period)
            try:
                normalized = normalize_value(value, scale) if value not in (None, "") else None
            except ValueError:
                normalized = None
            rows.append(
                {
                    "template_level": template_level,
                    "statement": record.get("Statement"),
                    "section": record.get("Section"),
                    "field_code": record.get("Field Code"),
                    "standard_field_name": record.get("Standard Field Name"),
                    "line_type": record.get("Line Type"),
                    "required_level": record.get("Required Level"),
                    "period": period,
                    "value": value,
                    "normalized_value": normalized,
                }
            )
    return pd.DataFrame(rows)


def get_period_columns(path: str | Path, sheet_name: str) -> list[str]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return [str(header) for header in headers if header not in INPUT_HEADERS and header is not None]


def _has_sufficient_required_data(df: pd.DataFrame, template_level: str) -> bool:
    if df.empty:
        return False
    required_inputs = {
        field.field_code
        for field in get_fields(template_level)
        if field.line_type == "Input" and field.required_level == REQUIRED
    }
    if not required_inputs:
        return False
    for code in required_inputs:
        subset = df[(df["field_code"] == code) & (df["normalized_value"].notna())]
        if subset.empty:
            return False
    return True


def detect_template_level(path: str | Path, profile: dict[str, Any] | None = None) -> str | None:
    profile = profile or read_company_profile(path)
    comprehensive = _read_input_sheet(path, SHEET_COMPREHENSIVE_INPUT, TEMPLATE_COMPREHENSIVE, profile)
    if _has_sufficient_required_data(comprehensive, TEMPLATE_COMPREHENSIVE):
        return TEMPLATE_COMPREHENSIVE
    essential = _read_input_sheet(path, SHEET_ESSENTIAL_INPUT, TEMPLATE_ESSENTIAL, profile)
    if _has_sufficient_required_data(essential, TEMPLATE_ESSENTIAL):
        return TEMPLATE_ESSENTIAL
    return None


def read_financial_data(path: str | Path, template_level: str | None = None) -> tuple[str | None, pd.DataFrame, dict[str, Any]]:
    profile = read_company_profile(path)
    selected = template_level or detect_template_level(path, profile)
    if selected == TEMPLATE_COMPREHENSIVE:
        sheet = SHEET_COMPREHENSIVE_INPUT
    elif selected == TEMPLATE_ESSENTIAL:
        sheet = SHEET_ESSENTIAL_INPUT
    else:
        return None, pd.DataFrame(), profile
    return selected, _read_input_sheet(path, sheet, selected, profile), profile


def dataframe_to_period_maps(df: pd.DataFrame) -> dict[str, dict[str, float]]:
    period_maps: dict[str, dict[str, float]] = {}
    if df.empty:
        return period_maps
    for period, group in df.dropna(subset=["normalized_value"]).groupby("period"):
        period_maps[str(period)] = dict(zip(group["field_code"], group["normalized_value"]))
    return period_maps
