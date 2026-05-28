"""Excel workbook generator for standardized input templates."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .field_codes import (
    CALCULATION_GUIDE_ROWS,
    INPUT_HEADERS,
    PROFILE_FIELDS,
    REPORTING_FREQUENCIES,
    REPORTING_SCALE_FACTORS,
    REQUIRED_SHEETS,
    SHEET_CALCULATION_GUIDE,
    SHEET_COMPANY_PROFILE,
    SHEET_COMPREHENSIVE_INPUT,
    SHEET_ESSENTIAL_INPUT,
    SHEET_FIELD_DICTIONARY,
    SHEET_INSTRUCTIONS,
    SHEET_VALIDATION_REPORT,
    TEMPLATE_COMPREHENSIVE,
    TEMPLATE_ESSENTIAL,
    TOOL_TITLE,
    get_fields,
)
from .utils import ensure_directory, project_root


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E2F0D9")
CHECK_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def build_period_labels(start_period: str = "FY2022", end_period: str = "FY2024") -> list[str]:
    start_year = int("".join(ch for ch in start_period if ch.isdigit()) or 2022)
    end_year = int("".join(ch for ch in end_period if ch.isdigit()) or start_year + 2)
    if end_year < start_year:
        end_year = start_year
    return [f"FY{year}" for year in range(start_year, end_year + 1)]


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.protection = Protection(locked=True)


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 42)


def _write_company_profile(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_COMPANY_PROFILE)
    ws.append(["Field", "Value"])
    defaults = {
        "Company Name": "Sample Private Company LLC",
        "Registration Number": "REG-0001",
        "Base Currency": "USD",
        "Country / Jurisdiction": "United Arab Emirates",
        "Reporting Frequency": "Annual",
        "Financial Year End": "31-Dec",
        "Analysis Start Period": "FY2022",
        "Analysis End Period": "FY2024",
        "Reporting Scale": "Millions",
        "Prepared By": "",
        "Date of Preparation": date.today().isoformat(),
    }
    for field in PROFILE_FIELDS:
        ws.append([field, defaults.get(field, "")])
    _style_header(ws)
    ws.freeze_panes = "A2"
    scale_dv = DataValidation(type="list", formula1=f'"{",".join(REPORTING_SCALE_FACTORS)}"', allow_blank=False)
    freq_dv = DataValidation(type="list", formula1=f'"{",".join(REPORTING_FREQUENCIES)}"', allow_blank=False)
    ws.add_data_validation(scale_dv)
    ws.add_data_validation(freq_dv)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == "Reporting Scale":
            scale_dv.add(ws.cell(row, 2))
        if ws.cell(row, 1).value == "Reporting Frequency":
            freq_dv.add(ws.cell(row, 2))
    _auto_width(ws)


def _write_input_sheet(wb: Workbook, sheet_name: str, template_level: str, periods: list[str]) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = list(INPUT_HEADERS) + periods
    ws.append(headers)
    _style_header(ws)
    ws.freeze_panes = "A2"
    for field in get_fields(template_level):
        row = [
            field.statement,
            field.section,
            field.field_code,
            field.standard_field_name,
            field.line_type,
            field.required_level,
            field.description,
        ] + [None for _ in periods]
        ws.append(row)
        fill = INPUT_FILL
        if field.line_type == "Calculated":
            fill = CALC_FILL
        elif field.line_type == "Check":
            fill = CHECK_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).protection = Protection(locked=col <= len(INPUT_HEADERS))
    for cell in ws[1]:
        cell.protection = Protection(locked=True)
    ws.protection.sheet = True
    ws.protection.password = "financials"
    _auto_width(ws)


def _write_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_INSTRUCTIONS)
    rows = [
        [TOOL_TITLE],
        ["Use either Essential_Input or Comprehensive_Input. If Comprehensive_Input has sufficient required data, it is processed first."],
        ["All calculations use Field Code values. Do not rename field codes."],
        ["Fill only the period value columns in the selected input sheet."],
        ["Set Reporting Scale in Company_Profile so uploaded values can be normalized internally."],
        ["This tool is for private-company analysis and does not use tickers, stock prices, market data, or public-company APIs."],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(size=16, bold=True)
    ws.column_dimensions["A"].width = 120


def _write_field_dictionary(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_FIELD_DICTIONARY)
    ws.append([
        "Template Level",
        "Statement",
        "Section",
        "Field Code",
        "Standard Field Name",
        "Line Type",
        "Required Level",
        "Formula",
        "Used In Ratios",
        "Description",
    ])
    _style_header(ws)
    for level in (TEMPLATE_ESSENTIAL, TEMPLATE_COMPREHENSIVE):
        for field in get_fields(level):
            ws.append([
                level,
                field.statement,
                field.section,
                field.field_code,
                field.standard_field_name,
                field.line_type,
                field.required_level,
                field.formula or "",
                ", ".join(field.used_in_ratios),
                field.description,
            ])
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_calculation_guide(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_CALCULATION_GUIDE)
    ws.append(["Category", "Calculation Step", "Formula Using Field Codes"])
    _style_header(ws)
    for row in CALCULATION_GUIDE_ROWS:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws["C2"].comment = Comment("All formulas intentionally use internal field codes.", "Codex")
    _auto_width(ws)


def _write_validation_report(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_VALIDATION_REPORT)
    ws.append([
        "severity",
        "template_level",
        "statement",
        "section",
        "field_code",
        "period",
        "expected_value",
        "actual_value",
        "difference",
        "message",
    ])
    _style_header(ws)
    ws.freeze_panes = "A2"
    _auto_width(ws)


def generate_template(output_path: str | Path | None = None, start_period: str = "FY2022", end_period: str = "FY2024") -> Path:
    output = Path(output_path) if output_path else project_root() / "templates" / "financial_statement_template.xlsx"
    ensure_directory(output.parent)
    wb = Workbook()
    wb.remove(wb.active)
    _write_instructions(wb)
    _write_company_profile(wb)
    periods = build_period_labels(start_period, end_period)
    _write_input_sheet(wb, SHEET_ESSENTIAL_INPUT, TEMPLATE_ESSENTIAL, periods)
    _write_input_sheet(wb, SHEET_COMPREHENSIVE_INPUT, TEMPLATE_COMPREHENSIVE, periods)
    _write_field_dictionary(wb)
    _write_calculation_guide(wb)
    _write_validation_report(wb)
    wb.save(output)
    return output


def _profile_value_row(ws, field_name: str) -> int:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == field_name:
            return row
    raise KeyError(field_name)


def _write_values(ws, values_by_code: dict[str, dict[str, float]]) -> None:
    headers = [cell.value for cell in ws[1]]
    period_columns = {header: idx + 1 for idx, header in enumerate(headers) if str(header).startswith("FY")}
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, 3).value
        if code in values_by_code:
            for period, value in values_by_code[code].items():
                if period in period_columns:
                    ws.cell(row, period_columns[period]).value = value


def generate_sample_data(output_path: str | Path | None = None) -> Path:
    output = Path(output_path) if output_path else project_root() / "sample_data" / "sample_private_company_data.xlsx"
    ensure_directory(output.parent)
    template_path = generate_template(project_root() / "templates" / "financial_statement_template.xlsx")
    wb = load_workbook(template_path)
    profile = wb[SHEET_COMPANY_PROFILE]
    profile.cell(_profile_value_row(profile, "Prepared By"), 2).value = "Finance Team"
    profile.cell(_profile_value_row(profile, "Company Name"), 2).value = "Atlas Components LLC"
    profile.cell(_profile_value_row(profile, "Registration Number"), 2).value = "ATLAS-7842"
    profile.cell(_profile_value_row(profile, "Reporting Scale"), 2).value = "Millions"
    values = {
        "cash_and_cash_equivalents": {"FY2022": 22, "FY2023": 26, "FY2024": 32},
        "short_term_investments": {"FY2022": 4, "FY2023": 5, "FY2024": 6},
        "accounts_receivable": {"FY2022": 38, "FY2023": 45, "FY2024": 52},
        "allowance_for_doubtful_debts": {"FY2022": 2, "FY2023": 2, "FY2024": 3},
        "net_accounts_receivable": {"FY2022": 36, "FY2023": 43, "FY2024": 49},
        "inventory": {"FY2022": 30, "FY2023": 36, "FY2024": 42},
        "prepaid_expenses": {"FY2022": 3, "FY2023": 4, "FY2024": 4},
        "other_current_assets": {"FY2022": 5, "FY2023": 6, "FY2024": 7},
        "total_current_assets": {"FY2022": 100, "FY2023": 120, "FY2024": 140},
        "property_plant_and_equipment": {"FY2022": 128, "FY2023": 145, "FY2024": 160},
        "accumulated_depreciation": {"FY2022": 28, "FY2023": 35, "FY2024": 43},
        "net_property_plant_and_equipment": {"FY2022": 100, "FY2023": 110, "FY2024": 117},
        "intangible_assets": {"FY2022": 8, "FY2023": 9, "FY2024": 10},
        "long_term_investments": {"FY2022": 6, "FY2023": 7, "FY2024": 8},
        "deferred_tax_assets": {"FY2022": 2, "FY2023": 2, "FY2024": 3},
        "other_non_current_assets": {"FY2022": 4, "FY2023": 4, "FY2024": 5},
        "total_non_current_assets": {"FY2022": 120, "FY2023": 132, "FY2024": 143},
        "total_assets": {"FY2022": 220, "FY2023": 252, "FY2024": 283},
        "accounts_payable": {"FY2022": 24, "FY2023": 28, "FY2024": 32},
        "accrued_expenses": {"FY2022": 8, "FY2023": 9, "FY2024": 10},
        "short_term_debt": {"FY2022": 12, "FY2023": 14, "FY2024": 15},
        "current_portion_of_long_term_debt": {"FY2022": 6, "FY2023": 7, "FY2024": 8},
        "tax_payable": {"FY2022": 3, "FY2023": 4, "FY2024": 5},
        "other_current_liabilities": {"FY2022": 7, "FY2023": 8, "FY2024": 10},
        "total_current_liabilities": {"FY2022": 60, "FY2023": 70, "FY2024": 80},
        "long_term_debt": {"FY2022": 45, "FY2023": 48, "FY2024": 50},
        "lease_liabilities": {"FY2022": 7, "FY2023": 8, "FY2024": 8},
        "deferred_tax_liabilities": {"FY2022": 3, "FY2023": 4, "FY2024": 5},
        "other_non_current_liabilities": {"FY2022": 5, "FY2023": 5, "FY2024": 7},
        "total_non_current_liabilities": {"FY2022": 60, "FY2023": 65, "FY2024": 70},
        "total_liabilities": {"FY2022": 120, "FY2023": 135, "FY2024": 150},
        "share_capital": {"FY2022": 50, "FY2023": 55, "FY2024": 60},
        "additional_paid_in_capital": {"FY2022": 10, "FY2023": 12, "FY2024": 13},
        "retained_earnings": {"FY2022": 32, "FY2023": 42, "FY2024": 52},
        "other_reserves": {"FY2022": 4, "FY2023": 5, "FY2024": 6},
        "other_equity": {"FY2022": 4, "FY2023": 3, "FY2024": 2},
        "total_equity": {"FY2022": 100, "FY2023": 117, "FY2024": 133},
        "total_liabilities_and_equity": {"FY2022": 220, "FY2023": 252, "FY2024": 283},
        "gross_revenue": {"FY2022": 250, "FY2023": 285, "FY2024": 325},
        "sales_returns_and_allowances": {"FY2022": 5, "FY2023": 7, "FY2024": 8},
        "revenue": {"FY2022": 245, "FY2023": 278, "FY2024": 317},
        "cost_of_goods_sold": {"FY2022": 150, "FY2023": 168, "FY2024": 190},
        "gross_profit": {"FY2022": 95, "FY2023": 110, "FY2024": 127},
        "selling_expenses": {"FY2022": 18, "FY2023": 20, "FY2024": 23},
        "general_and_administrative_expenses": {"FY2022": 24, "FY2023": 26, "FY2024": 29},
        "research_and_development_expenses": {"FY2022": 4, "FY2023": 5, "FY2024": 6},
        "other_operating_expenses": {"FY2022": 3, "FY2023": 4, "FY2024": 5},
        "total_operating_expenses_excluding_depreciation_and_amortization": {"FY2022": 49, "FY2023": 55, "FY2024": 63},
        "ebitda": {"FY2022": 46, "FY2023": 55, "FY2024": 64},
        "depreciation_and_amortization": {"FY2022": 12, "FY2023": 14, "FY2024": 16},
        "operating_profit": {"FY2022": 34, "FY2023": 41, "FY2024": 48},
        "interest_income": {"FY2022": 1, "FY2023": 1, "FY2024": 2},
        "finance_costs": {"FY2022": 7, "FY2023": 8, "FY2024": 9},
        "other_income": {"FY2022": 2, "FY2023": 2, "FY2024": 3},
        "other_expenses": {"FY2022": 1, "FY2023": 1, "FY2024": 2},
        "profit_before_tax": {"FY2022": 29, "FY2023": 35, "FY2024": 42},
        "tax_expense": {"FY2022": 7, "FY2023": 9, "FY2024": 10},
        "net_profit": {"FY2022": 22, "FY2023": 26, "FY2024": 32},
        "net_cash_flow_from_operating_activities": {"FY2022": 30, "FY2023": 36, "FY2024": 42},
        "net_cash_flow_from_investing_activities": {"FY2022": -18, "FY2023": -20, "FY2024": -21},
        "net_cash_flow_from_financing_activities": {"FY2022": 5, "FY2023": -2, "FY2024": -1},
        "net_change_in_cash": {"FY2022": 17, "FY2023": 14, "FY2024": 20},
        "opening_cash_balance": {"FY2022": 5, "FY2023": 12, "FY2024": 12},
        "closing_cash_balance": {"FY2022": 22, "FY2023": 26, "FY2024": 32},
    }
    _write_values(wb[SHEET_COMPREHENSIVE_INPUT], values)
    wb.save(output)
    return output


if __name__ == "__main__":
    generate_template()
    generate_sample_data()
