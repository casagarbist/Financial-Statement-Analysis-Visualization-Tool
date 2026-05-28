from openpyxl import load_workbook

from src.field_codes import SHEET_COMPANY_PROFILE, SHEET_COMPREHENSIVE_INPUT
from src.template_generator import generate_sample_data
from src.validation import validate_workbook


def _set_profile_value(path, field, value):
    wb = load_workbook(path)
    ws = wb[SHEET_COMPANY_PROFILE]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == field:
            ws.cell(row, 2).value = value
            break
    wb.save(path)


def _set_value(path, code, period, value):
    wb = load_workbook(path)
    ws = wb[SHEET_COMPREHENSIVE_INPUT]
    headers = [cell.value for cell in ws[1]]
    col = headers.index(period) + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 3).value == code:
            ws.cell(row, col).value = value
            break
    wb.save(path)


def test_missing_company_profile_field(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    _set_profile_value(path, "Company Name", None)
    issues, _ = validate_workbook(path)
    assert any(issue.severity == "Error" and "Company profile" in issue.message for issue in issues)


def test_invalid_reporting_scale(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    _set_profile_value(path, "Reporting Scale", "MegaUnits")
    issues, _ = validate_workbook(path)
    assert any(issue.field_code == "Reporting Scale" and issue.severity == "Error" for issue in issues)


def test_missing_required_essential_field(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    for code in ["gross_revenue", "revenue"]:
        _set_value(path, code, "FY2024", None)
    issues, _ = validate_workbook(path)
    assert any(issue.severity == "Error" and issue.field_code == "gross_revenue" for issue in issues)


def test_missing_optional_comprehensive_field_creates_warning(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    _set_value(path, "short_term_investments", "FY2024", None)
    issues, _ = validate_workbook(path)
    assert any(issue.severity == "Warning" and issue.field_code == "short_term_investments" for issue in issues)


def test_balance_sheet_mismatch(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    _set_value(path, "total_assets", "FY2024", 999)
    issues, _ = validate_workbook(path)
    assert any("Balance Sheet" in issue.message for issue in issues)


def test_profit_and_loss_mismatch(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    _set_value(path, "net_profit", "FY2024", 999)
    issues, _ = validate_workbook(path)
    assert any(issue.statement == "Profit & Loss" and issue.field_code == "net_profit" for issue in issues)


def test_cash_flow_mismatch(tmp_path):
    path = generate_sample_data(tmp_path / "sample.xlsx")
    _set_value(path, "closing_cash_balance", "FY2024", 999)
    issues, _ = validate_workbook(path)
    assert any(issue.statement == "Cash Flow" for issue in issues)
