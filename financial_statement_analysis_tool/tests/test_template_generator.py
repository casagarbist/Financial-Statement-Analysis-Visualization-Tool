from openpyxl import load_workbook

from src.field_codes import REQUIRED_SHEETS, SHEET_COMPANY_PROFILE
from src.template_generator import generate_template


def test_template_generator_creates_required_sheets(tmp_path):
    path = generate_template(tmp_path / "template.xlsx")
    wb = load_workbook(path)
    assert wb.sheetnames == list(REQUIRED_SHEETS)
    assert wb[SHEET_COMPANY_PROFILE]["A1"].value == "Field"


def test_input_sheets_have_frozen_headers(tmp_path):
    path = generate_template(tmp_path / "template.xlsx")
    wb = load_workbook(path)
    assert wb["Essential_Input"].freeze_panes == "A2"
    assert wb["Comprehensive_Input"].freeze_panes == "A2"
